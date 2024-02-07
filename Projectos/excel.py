from time import sleep
from colorama import init, Fore, Style
from openpyxl import Workbook, load_workbook

# Localização do ficheiro Excel
location = r'E:\Biblioteca.xlsx'

# Tentar carregar o livro de Excel; se não existir, criar um novo
try:
    workbook = load_workbook(location)
except FileNotFoundError:
    workbook = Workbook()
    workbook.save(location)

# Criar ou carregar a folha 'Livros'
sheet = workbook.create_sheet('Livros', 0) if 'Livros' not in workbook.sheetnames else workbook['Livros']

# Inicializar o colorama
init(autoreset=True)

# Conjuntos para armazenar títulos e ISBNs registados
titulos_registados = list(row[0].value.lower() for row in sheet.iter_rows(min_row=2, max_col=1))
isbns_registados = list(row[2].value for row in sheet.iter_rows(min_row=2, max_col=3))


def menu_principal():
    """
    Função que exibe o menu principal.
    """
    menu_word = "=== Menu ==="
    print("\n" + Fore.CYAN + Style.BRIGHT + menu_word)
    print(Fore.YELLOW + "[1] Inserir Livros")
    print(Fore.YELLOW + "[2] Listar Livros")
    print(Fore.YELLOW + "[3] Pesquisar Livros")
    print(Fore.YELLOW + "[4] Remover Livros")
    print(Fore.RED + "[5] Sair")
    print(Fore.CYAN + Style.BRIGHT + "=" * len(menu_word))


def validador_isbn(isbn):
    """
    Função que verifica se um ISBN é válido.
    :param isbn: O ISBN a verificar
    :return: True se o ISBN for válido, False caso contrário
    """
    print(f"A verificar ISBN: {isbn}")
    print(f"ISBNs registados: {isbns_registados}")
    return len(isbn) == 13 and isbn.isnumeric() and isbn not in isbns_registados


def isbn_valido():
    """
    Função que solicita e valida um ISBN.
    :return: O ISBN válido
    """
    while True:
        isbn = input('Digite o ISBN do livro: ')
        if validador_isbn(isbn):
            return isbn
        else:
            print('ISBN inválido ou já registado. Tente novamente.')


def titulo_valido():
    """
    Função que solicita e valida um título.
    :return: O título válido
    """
    title = input("Digite o título do livro: ").strip().lower()
    while title in titulos_registados:
        print('Este livro já está registado')
        title = input("Digite o título do livro: ").strip().lower()
    return title


def inserir_menu():
    """
    Função principal para inserir livros.
    """
    while True:
        # Carregar dados existentes da folha Excel para a lista biblioteca
        sheet = workbook['Livros']
        biblioteca = set(row[0].value.lower() for row in sheet.iter_rows(min_row=2, max_col=1))

        # Definir um dicionário para cada livro
        livros = dict()

        # Solicitar e validar título
        titulo = titulo_valido()
        livros['Título'] = titulo
        titulos_registados.append(titulo)  # Adicionar o título ao conjunto

        # Solicitar e validar autores
        lista_autores = list()
        while True:
            autor = [input("Digite o autor do livro: ")]
            lista_autores.append(autor)
            continuar = input("O livro tem mais autores?: [S/N] ").strip().lower()
            if continuar == "n":
                break
        livros["Autores"] = lista_autores

        # Solicitar e validar ISBN
        isbn = isbn_valido()
        livros['ISBN'] = isbn

        # Atualizar o conjunto de ISBNs registados
        isbns_registados.append(isbn)

        # Solicitar e validar ano de publicação
        ano_publicacao = input('Digite o ano de publicação: ')
        while len(ano_publicacao) != 4 or not ano_publicacao.isnumeric():
            ano_publicacao = input('Digite corretamente o ano de publicação do livro: ')
        else:
            livros['Ano de publicação'] = ano_publicacao

        # Solicitar editora
        livros['Editora'] = input('Digite a editora do livro: ').strip()

        # Solicitar e validar géneros/categorias
        categorias = list()
        while True:
            cat = [input("Digite a categoria do livro: ")]
            categorias.append(cat)
            continuar = input("O livro pode ser identificado com outra categoria?: [S/N] ").strip().lower()
            if continuar == "n":
                livros["Categorias"] = categorias
                break

        # Definir disponibilidade do livro
        status = 'Disponível'
        livros['Status'] = status

        # Criar uma linha para adicionar à folha
        row = [livros['Título'], ', '.join([' '.join(keys) for keys in livros['Autores']]),
               livros['ISBN'], livros.get('Ano de publicação', ''), livros.get('Editora', ''),
               ', '.join([' '.join(keys) for keys in livros.get('Categorias', [])]), livros['Status']]

        # Adicionar registo
        sheet.append(row)
        print('Livro registado')
        workbook.save(location)
        terminar = input('Deseja registar mais algum livro? [S/N] ').strip().upper()
        if terminar == 'N':
            break



def listar_menu(biblio):
    for livro in biblio:
        for key in livro:
            if isinstance(livro[key], list):
                formated_keys = ', '.join([' '.join(keys) for keys in livro[key]])
                print(Fore.BLUE + f"{key}: {Fore.WHITE + formated_keys}")
            else:
                print(Fore.BLUE + f"{key}: {Fore.WHITE + livro[key]}")
        print(Style.RESET_ALL)  # Reset text color/style
        print(Fore.GREEN + "*=*" * 15)
        print()


def remover_livro(ls, op_remover):
    temp_livros = []
    if len(ls) == 0:
        print("Não existe livros na biblioteca")
    else:
        print(f"{op_remover.upper()}S disponiveis: ", end=" ")
        for i, livro in enumerate(ls):
            print(livro[op_remover], end="")
            if i < len(ls) - 1:
                print(", ", end="")
        valor_remover = input(f"\nQual o {op_remover} do livro a remover?: ") if op_remover == "Título" else (
            input(f"\n{op_remover.upper()} do livro a remover: "))

        for livro in ls:
            if livro[op_remover] != valor_remover:
                temp_livros.append(livro)
    return temp_livros


def remover_menu(li):
    livros_remover = []
    while True:
        opcao_remover = input("Pretende remover livros a partir do que? [ISBN ou Título]: ").strip()
        while opcao_remover not in ["ISBN", "Título"]:
            print("Opção inválida, escreve ISBN ou Título")
            opcao_remover = input("Pretende remover livros a partir do que? [ISBN ou Título]: ").strip()
        livros_remover = remover_livro(li, opcao_remover)

        opcao_remover = input("Pretende remover mais? [s/n]: ").lower().strip()
        if opcao_remover == "n":
            break
        if len(livros_remover) == 0:
            print("Já não existe livros para remover")
            sleep(1)
            print("A sair do menu de eliminação de livros", end=" ")
            sleep(0.5)
            print(".", end=" ")
            sleep(0.5)
            print(".", end=" ")
            sleep(0.5)
            print(".")
            break
    return livros_remover


'''biblioteca = [
    {
        'Título': 'ola', 'Autores': [['olad'], ['ola123']], 'ISBN': '1234567890123', 'Ano de publicação': '1321',
        'Editora': 'ola',
        'Categorias': [['a']], 'Status': 'Disponível'
    },
    {
        'Título': 'ola123', 'Autores': [['oladddd'], ['ola123444'], ['autorrr']], 'ISBN': '1234567890125',
        'Ano de publicação': '3213', 'Editora': 'ola',
        'Categorias': [['a']], 'Status': 'Disponível'
    }
]'''

while True:
    menu_principal()
    choice = input("Escolha um opção (1-5): ").strip()
    if choice == "1":
        inserir_menu()
    elif choice == "2":
        listar_menu(biblioteca)
    elif choice == "3":
        break
    elif choice == "4":
        biblioteca = remover_menu(biblioteca)
    elif choice == "5":
        print("Programa a terminar", end=" ")
        sleep(0.5)
        print(".", end=" ")
        sleep(0.5)
        print(".", end=" ")
        sleep(0.5)
        print(".")
        break
    else:
        print("Escolha inválida. Escolha uma opção entre 1-5.")
